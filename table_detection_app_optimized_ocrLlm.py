import os
import tempfile
from io import BytesIO

import cv2
import data_extraction_ocr_llm
import numpy as np
import pandas as pd
import streamlit as st
import table_detection
from data_extraction_ocr_llm import OCREngine
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
from SwinSR_Model.swin2sr_net import ImageResolutionUpScaler

# Title of the app
st.title("Table Detection and Data Extraction Demo Application Optimized OCR-LLM")

# Instructions
st.write("Upload an image (JPG or PNG format) to process and extract the table data.")

# File uploader
uploaded_file = st.file_uploader("Choose an image...", type=["jpg", "png"])


# Initialize the detector model and cache it
@st.cache_resource
def load_detector_model(cfg_path, model_path):
    return table_detection.initialize_detector_model(str(model_path), str(cfg_path))


# Load the detector model
cfg_path = "./All_X152.yaml"
model_path = "./model_final.pth"
predictor = load_detector_model(cfg_path, model_path)

sr_model_pth = "./SwinSR_Model/models"
upscaler = ImageResolutionUpScaler(
    sr_model_pth + "/" + "Swin2SR_RealworldSR_X4_64_BSRGAN_PSNR.pth"
)  # Model path

ocr_engine = OCREngine()


def create_dataframe(data_dict):
    # Convert the dictionary to DataFrame
    df = pd.DataFrame(data_dict).T
    return df


def create_excel_from_dataframe(df):
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    # Write headers: data keys as the first row
    headers = list(df.index)
    for row_idx, header in enumerate(headers, start=1):
        ws.cell(row=row_idx, column=1, value=header)

    # Write DataFrame values row by row, handling lists by writing each item in separate cells
    for r_idx, (index, row) in enumerate(df.iterrows(), start=1):
        c_idx = 2  # Start column index at 1
        for value in row:
            if isinstance(value, list):
                for item in value:
                    ws.cell(row=r_idx, column=c_idx, value=item)
                    c_idx += 1  # Move to the next column

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Reset the file pointer to the beginning
    return output


# # Function to convert DataFrame to Excel and provide download link
# def to_excel(df, file):
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, headers=False, sheet_name="Extracted Data")
#     processed_data = output.getvalue()
#     return processed_data
#     # data = df.to_excel(
#     #     file, index=False, engine="openpyxl", sheet_name="Extracted Data"
#     # )
#     # return data
# Initialize session state for download flag
if "downloaded" not in st.session_state:
    st.session_state.downloaded = False

if uploaded_file is not None:
    # Open and display the uploaded image
    # Get the file name
    file_name = uploaded_file.name
    st.write(f"File name: {file_name}")

    # Read the uploaded image using OpenCV
    image_bytes = uploaded_file.read()  # Read image as bytes
    image_array = np.frombuffer(image_bytes, np.uint8)  # Convert bytes to numpy array
    image_bgr = cv2.imdecode(
        image_array, cv2.IMREAD_COLOR
    )  # Decode image array using OpenCV
    # Convert BGR image to RGB format
    image = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
    imgH, imgW, _ = image.shape
    is_img_large = imgW > 1000 and imgH > 900

    # Set initial thresholds based on image size
    if is_img_large:
        score_thresh = 0.07  # 0.08,0.09
        nms_thresh = 0.58  # 0.58, 0.5
        predictor = table_detection.update_predictor_thresholds(
            predictor, score_thresh, nms_thresh
        )

    # print("image shape", image.shape)

    # Process the image
    table_list, table_coords, images = table_detection.make_prediction(image, predictor)
    # Make predictions on the new images obtained from the first run three times
    if is_img_large:
        for i in range(3):
            if i == 1:
                score_thresh = 0.13  # 0.13, 0.1, 0.2
                nms_thresh = 0.35  # 0.34, 0.37, 0.41
                predictor = table_detection.update_predictor_thresholds(
                    predictor, score_thresh, nms_thresh
                )

            new_table_list = []
            new_table_coords = []
            new_images = []  # Create a new list to store new predictions
            for img in images:
                table_list_pred, table_coords_pred, img_predictions = (
                    table_detection.make_prediction(img, predictor)
                )
                new_images.extend(img_predictions)  # Add new predictions to the list
                # new_table_list.extend()
            # Update 'images' with the new predictions for the next iteration
            images = new_images
    # Filter images with only table data
    # extracted_images = table_detection.filter_images_with_tables(images)
    # Display the uploaded image
    st.image(
        image,
        caption="Uploaded Image",
        use_column_width=True,
    )
    # print(len(extracted_images))
    # Display each processed image
    extracted_data_list = []
    for i, img in enumerate(images):
        # print(type(img))
        # Convert the processed image back to PIL format
        processed_image = Image.fromarray(img)
        try:
            processed_image = Image.fromarray(upscaler.imageSuperResolution(img))
        except RuntimeError as e:
            if "out of memory" in str(e).lower():
                st.write(
                    "CUDA out of memory error encountered. Processing on CPU instead."
                )
                # processed_image = Image.fromarray(img)
            else:
                raise e

        # Display the processed image
        st.image(
            processed_image, caption=f"Extracted Table {i + 1}", use_column_width=True
        )
        extracted_data = data_extraction_ocr_llm.table_data_extraction(
            processed_image, ocr_engine
        )
        df = create_dataframe([extracted_data])
        # combined_df = create_dataframe([extracted_data])
        # # Convert DataFrame to Excel
        # excel_data = to_excel(combined_df)
        file = f"extracted_data_{file_name}_{i + 1}.xlsx"
        excel_data = create_excel_from_dataframe(df)
        # extracted_data_list.append(extracted_data)
        st.json(extracted_data)
        # Create a DataFrame

        # Add a download button for each extracted data
        download_label = f"Download Extracted Data {i + 1} as Excel"
        # st.markdown(
        #     f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_data}" download="extracted_data_{file_name}_{i}.xlsx">{download_label}</a>',
        #     unsafe_allow_html=True,
        # )
        st.download_button(
            label=download_label,
            data=excel_data,
            file_name=file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Function to create a DataFrame from the extracted data
    # print("extracted data list", extracted_data_list)
